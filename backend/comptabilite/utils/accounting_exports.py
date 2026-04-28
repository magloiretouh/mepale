"""
MEPALE ERP — Exports comptabilité
  - export_entries_excel : liste des écritures en .xlsx (openpyxl)
  - export_report_pdf    : compte de résultats en PDF A4 (reportlab)
"""

from io import BytesIO


# ---------------------------------------------------------------------------
# Excel — openpyxl
# ---------------------------------------------------------------------------

def export_entries_excel(entries, date_from, date_to) -> bytes:
    """
    Génère un fichier .xlsx listant les écritures fournies.

    Paramètres
    ----------
    entries   : queryset ou liste d'EcritureComptable
    date_from : date (début de période)
    date_to   : date (fin de période)

    Retourne
    --------
    bytes du fichier .xlsx
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ACCENT      = '00A88C'
    ACCENT_LIGHT= 'E6FAF6'
    RED         = 'EF4444'
    RED_LIGHT   = 'FEF2F2'
    HEADER_TEXT = 'FFFFFF'
    THIN        = Side(style='thin', color='D1D5DB')
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Écritures comptables'

    # ── Titre ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value      = f"Écritures comptables — {date_from.strftime('%d/%m/%Y')} au {date_to.strftime('%d/%m/%Y')}"
    title_cell.font       = Font(bold=True, size=13, color=HEADER_TEXT)
    title_cell.fill       = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type='solid')
    title_cell.alignment  = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── En-têtes colonnes ──────────────────────────────────────────────────
    headers = ['Date', 'Type', 'Catégorie', 'Libellé', 'Montant (FCFA)', 'Source', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font      = Font(bold=True, color=HEADER_TEXT, size=10)
        cell.fill      = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = BORDER
    ws.row_dimensions[2].height = 22

    # ── Données ────────────────────────────────────────────────────────────
    for row_idx, entry in enumerate(entries, 3):
        is_income = entry.type == 'income'
        row_fill  = PatternFill(
            start_color=ACCENT_LIGHT if is_income else RED_LIGHT,
            end_color  =ACCENT_LIGHT if is_income else RED_LIGHT,
            fill_type  ='solid',
        ) if row_idx % 2 == 0 else None

        values = [
            entry.date.strftime('%d/%m/%Y'),
            'Recette' if is_income else 'Charge',
            entry.category.name if entry.category else 'Non catégorisé',
            entry.label,
            float(entry.amount),
            'Manuel' if entry.source == 'manual' else 'Auto',
            entry.notes or '',
        ]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = BORDER
            cell.alignment = Alignment(vertical='center')
            if row_fill:
                cell.fill = row_fill
            # Montant en gras + couleur
            if col == 5:
                cell.font      = Font(bold=True, color=ACCENT if is_income else RED)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # ── Largeur auto colonnes ──────────────────────────────────────────────
    col_widths = [12, 10, 28, 42, 18, 8, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header rows
    ws.freeze_panes = 'A3'

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF — reportlab
# ---------------------------------------------------------------------------

def export_report_pdf(data, date_from, date_to) -> bytes:
    """
    Génère un PDF A4 du compte de résultats.

    Paramètres
    ----------
    data      : dict retourné par _build_report_data()
    date_from : date
    date_to   : date

    Retourne
    --------
    bytes du fichier PDF
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    ACCENT      = colors.HexColor('#00A88C')
    ACCENT_LIGHT= colors.HexColor('#E6FAF6')
    RED         = colors.HexColor('#EF4444')
    RED_LIGHT   = colors.HexColor('#FEF2F2')
    GREY_LIGHT  = colors.HexColor('#F9FAFB')
    GREY_BORDER = colors.HexColor('#E5E7EB')
    WHITE       = colors.white
    DARK        = colors.HexColor('#111827')

    PAGE_W = A4[0] - 4 * cm  # largeur utile
    COL_CAT   = PAGE_W * 0.65
    COL_TOTAL = PAGE_W * 0.35

    buffer = BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles   = getSampleStyleSheet()
    story    = []

    # ── Styles personnalisés ───────────────────────────────────────────────
    style_title = ParagraphStyle(
        'title', fontSize=18, fontName='Helvetica-Bold',
        textColor=DARK, spaceAfter=4,
    )
    style_subtitle = ParagraphStyle(
        'subtitle', fontSize=10, fontName='Helvetica',
        textColor=colors.HexColor('#6B7280'), spaceAfter=16,
    )
    style_section = ParagraphStyle(
        'section', fontSize=11, fontName='Helvetica-Bold',
        textColor=DARK, spaceBefore=12, spaceAfter=6,
    )

    # ── En-tête ────────────────────────────────────────────────────────────
    story.append(Paragraph('Compte de Résultats', style_title))
    story.append(Paragraph(
        f"Période : {date_from.strftime('%d/%m/%Y')} — {date_to.strftime('%d/%m/%Y')}",
        style_subtitle,
    ))
    story.append(HRFlowable(width='100%', thickness=2, color=ACCENT, spaceAfter=16))

    # ── Fonction tableau section ──────────────────────────────────────────
    def _section_table(rows, total, header_color, total_bg, total_text):
        table_data = [['Catégorie', 'Montant (FCFA)']]

        if rows:
            for row in rows:
                table_data.append([
                    row['category'],
                    f"{row['total']:,.0f}",
                ])
        else:
            table_data.append(['Aucune écriture sur cette période.', '—'])

        table_data.append(['TOTAL', f"{total:,.0f}"])

        t = Table(table_data, colWidths=[COL_CAT, COL_TOTAL])

        style = TableStyle([
            # En-tête
            ('BACKGROUND',   (0, 0), (-1, 0), header_color),
            ('TEXTCOLOR',    (0, 0), (-1, 0), WHITE),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 9),
            ('TOPPADDING',   (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING',(0, 0), (-1, 0), 8),

            # Corps
            ('FONTNAME',  (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE',  (0, 1), (-1, -2), 9),
            ('TEXTCOLOR', (0, 1), (-1, -2), DARK),
            ('TOPPADDING',    (0, 1), (-1, -2), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 6),

            # Alternance lignes
            *[
                ('BACKGROUND', (0, i), (-1, i), GREY_LIGHT)
                for i in range(2, len(table_data) - 1, 2)
            ],

            # Total
            ('BACKGROUND', (0, -1), (-1, -1), total_bg),
            ('TEXTCOLOR',  (0, -1), (-1, -1), total_text),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, -1), (-1, -1), 10),
            ('TOPPADDING',    (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),

            # Alignements
            ('ALIGN',  (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN',  (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # Bordures
            ('GRID',      (0, 0), (-1, -1), 0.5, GREY_BORDER),
            ('LINEBELOW', (0, 0), (-1, 0),  1,   header_color),
        ])
        t.setStyle(style)
        return t

    # ── Section Recettes ──────────────────────────────────────────────────
    story.append(Paragraph('RECETTES', style_section))
    story.append(_section_table(
        data['income']['rows'],
        data['income']['total'],
        header_color=ACCENT,
        total_bg=ACCENT_LIGHT,
        total_text=ACCENT,
    ))
    story.append(Spacer(1, 20))

    # ── Section Charges ───────────────────────────────────────────────────
    story.append(Paragraph('CHARGES', style_section))
    story.append(_section_table(
        data['expense']['rows'],
        data['expense']['total'],
        header_color=RED,
        total_bg=RED_LIGHT,
        total_text=RED,
    ))
    story.append(Spacer(1, 24))

    # ── Résultat net ─────────────────────────────────────────────────────
    net       = data['net_result']
    is_profit = net >= 0
    net_color = ACCENT if is_profit else RED
    net_bg    = ACCENT_LIGHT if is_profit else RED_LIGHT
    net_label = 'BÉNÉFICE NET' if is_profit else 'PERTE NETTE'

    net_table = Table(
        [[net_label, f"{abs(net):,.0f} FCFA"]],
        colWidths=[COL_CAT, COL_TOTAL],
    )
    net_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), net_bg),
        ('TEXTCOLOR',     (0, 0), (-1, -1), net_color),
        ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 13),
        ('TOPPADDING',    (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('ALIGN',         (1, 0), (1, -1),  'RIGHT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX',           (0, 0), (-1, -1), 2, net_color),
    ]))
    story.append(net_table)

    doc.build(story)
    return buffer.getvalue()
